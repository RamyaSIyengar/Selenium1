import sys
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pytest
import openpyxl
from selenium.webdriver.common.by import By


def get_data():
    # return [
    #     ['Admin', 'admin123'],
    #     ['Admin', 'admin1234'],
    #     # ('Admin12', 'admin123'),
    #     # ('Admin123', 'admin1234')
    # ]

    workbook = openpyxl.load_workbook("C:\\Users\\soundarr\\Documents\\data.xlsx")
    sheet = workbook["Sheet4"]
    rows = sheet.max_row
    cols = sheet.max_column
    mainList = []

    for i in range(2, rows + 1):
        dataList = []
        for j in range(1, cols + 1):
            data = sheet.cell(i, j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    print(mainList)
    return mainList


# @pytest.mark.usefixtures("setup")
class TestLogin:


    @pytest.mark.parametrize("username, password", get_data())
    @pytest.mark.smoke
    def test_userReg(self, username, password, setup):
        self.driver = setup
        try:
            self.usr =WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.NAME, "username")))
            self.passwd =WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.NAME, "password")))
        except TimeoutError:
            pass
        self.usr.send_keys(username)
        self.passwd.send_keys(password)
        self.btn = self.driver.find_element(By.TAG_NAME, "button")
        self.driver.execute_script("arguments[0].scrollIntoView(true)", self.btn)
        self.btn.click()

        self.dashbd = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "//h6[normalize-space()='Dashboard']")))
        assert "Dashboard" in self.dashbd.text
        print("In home page")

    # @pytest.mark.skip(reason="Not ready for testing")
    # @pytest.mark.smoke
    # def test_doLogin(self):
    #         assert "OrangeHRM" in self.driver.title
    #         print("Executing Login")
    #
    # @pytest.mark.skip(reason="Not ready for testing")
    # @pytest.mark.regression
    # def test_search(self):
    #         assert "OrangeHRM" in self.driver.title
    #         print("Executing Login")
    #
    # @pytest.mark.skip(reason="Not ready for testing")
    # def test_skip(self):
    #     print("skipping this test by using builtin marker @pytest.mark.skip")

# Skips the test if a certain condition is met
#     @pytest.mark.skipif(sys.platform == "win32", reason="Does not run on Windows")
#     def test_only_on_linux(self):
#         assert True

    """Marks a test as expected to fail.If the test fails, it won \'t be considered a failure, '
     'and if it unexpectedly passes, it will be reported as an "XFAIL."""

    # @pytest.mark.xfail(reason="Bug in the code")
    # def test_bug(self):
    #     assert False

