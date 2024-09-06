import openpyxl

file = "C:\\Users\\soundarr\\Documents\\data.xlsx"


# # # Reading data from XL
# # # File > workbook > sheets > Rows > cells
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet4"]
#
# rows = sheet.max_row  # no of rows in Excel table
# col = sheet.max_column  # no of rows in Excel table
# print(rows, col)
#
# data = []
# for r in range(2, rows+1):
#     for c in range(1, col+1):
#         print(sheet.cell(r, c).value, end='      ')
#         data.append(sheet.cell(r, 1).value)
#     print()
# print(data)


# Writing to Excel sheet

# 1. same data on each cell
# workbook = openpyxl.load_workbook(filename=file)
# sheet = workbook["Sheet3"]
#
# print(sheet)
#
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "welcome"
#
# workbook.save(filename=file)

# 2. writing multiple data on cells
# workbook = openpyxl.load_workbook(filename=file)
# sheet = workbook["Sheet4"]
#
# print(sheet)
#
# sheet.cell(1, 1).value = "welcome"
# sheet.cell(1, 2).value = "Back"
# sheet.cell(1, 3).value = "to"
# sheet.cell(2, 1).value = "python"
# sheet.cell(2, 2).value = "selenium"
# sheet.cell(2, 3).value = "pytest"
#
# workbook.save(filename=file)
#


